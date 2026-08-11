"""Regenerate Leave-Payroll-Implementation-Tracker.xlsx from CSV sheets."""
from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
CSV_DIR = ROOT / "leave-payroll-implementation-tracker-csv"
OUT = ROOT / "Leave-Payroll-Implementation-Tracker.xlsx"

HEADERS = [
    "Section",
    "Sub-Section",
    "Item / Task",
    "Layer",
    "Detail / Expected Result",
    "Status",
    "Assigned Dev",
    "Priority",
    "Sprint / Phase",
    "Notes",
]

FILES = [
    ("00 Overview", "00_Overview.csv"),
    ("01 Leave-Clash-Swap", "01_Leave-Clash-Swap.csv"),
    ("02 Attendance-Holidays-Payroll", "02_Attendance-Holidays-Payroll.csv"),
    ("03 Frontend-UAT", "03_Frontend-UAT.csv"),
]

header_font = Font(bold=True)
header_fill = PatternFill("solid", fgColor="D9E1F2")
section_fill = PatternFill("solid", fgColor="F2F2F2")
section_font = Font(bold=True)
wrap = Alignment(wrap_text=True, vertical="top")


def is_section_row(row: list[str]) -> bool:
    section = row[0] if row else ""
    sub = row[1] if len(row) > 1 else ""
    item = row[2] if len(row) > 2 else ""
    return bool(section) and not sub and not item


def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, filename in FILES:
        ws = wb.create_sheet(sheet_name)
        for col, header in enumerate(HEADERS, start=1):
            cell = ws.cell(1, col, header)
            cell.font = header_font
            cell.fill = header_fill

        path = CSV_DIR / filename
        with path.open(encoding="utf-8", newline="") as handle:
            rows = list(csv.reader(handle))

        for excel_row, raw in enumerate(rows[1:], start=2):
            row = (raw + [""] * 10)[:10]
            for col, value in enumerate(row, start=1):
                cell = ws.cell(excel_row, col, value)
                cell.alignment = wrap
            if is_section_row(row):
                for col in range(1, 11):
                    ws.cell(excel_row, col).fill = section_fill
                ws.cell(excel_row, 1).font = section_font

        widths = [18, 16, 42, 12, 55, 14, 14, 12, 16, 48]
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width
        ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"
        ws.freeze_panes = "A2"

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
